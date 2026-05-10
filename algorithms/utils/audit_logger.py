from datetime import datetime
from typing import List, Dict, Any, Optional

class PlanningAuditLogger:
    def __init__(self):
        self.logs: List[Dict[str, Any]] = []
        self._start_time = datetime.now()

    def add_log(self, time: datetime, action: str, details: str, step: str = "GENERAL"):
        self.logs.append({
            "timestamp": time,
            "step": step,
            "action": action,
            "details": details
        })

    def generate_report_xlsx(self, file_path: str, summary_data: Optional[Dict[str, Any]] = None):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = openpyxl.Workbook()
            
            # --- STİLLER ---
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # --- 1. SAYFA: GENEL ÖZET (EĞER VARSA) ---
            if summary_data:
                ws_sum = wb.active
                ws_sum.title = "Genel Özet"
                
                sum_headers = ["Ürün Tipi", "Ürün Adı", "Dönemlik Hedef", "Excel'den Gelen", "Planlanan Adet", "Kalan"]
                ws_sum.append(["📊 ÜRETİM PLANLAMA ÖZET RAPORU"])
                ws_sum.merge_cells('A1:F1')
                ws_sum['A1'].font = Font(bold=True, size=14)
                ws_sum['A1'].alignment = center_align
                
                ws_sum.append([]) # Boş satır
                ws_sum.append(sum_headers)
                
                for cell in ws_sum[3]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = border

                for row in summary_data.get("products", []):
                    ws_sum.append([
                        row.get("type", "-"),
                        row.get("name", "-"),
                        row.get("target", 0),
                        row.get("excel_count", 0),
                        row.get("scheduled", 0),
                        row.get("remaining", 0)
                    ])
                
                # Özet tablosu stilleri
                for row_idx in range(4, ws_sum.max_row + 1):
                    for col_idx in range(1, 7):
                        cell = ws_sum.cell(row=row_idx, column=col_idx)
                        cell.border = border
                        cell.alignment = center_align

                ws_sum.column_dimensions['A'].width = 15
                ws_sum.column_dimensions['B'].width = 25
                for col in ['C', 'D', 'E', 'F']: ws_sum.column_dimensions[col].width = 15
                
                ws_analysis = wb.create_sheet("Karar Analizi")
            else:
                ws_analysis = wb.active
                ws_analysis.title = "Karar Analizi"

            # --- 2. SAYFA: KARAR ANALİZİ ---
            headers = ["Sıra", "Zaman", "İstasyon", "İşlem / Karar", "Seçilen Ürün", "Karar Gerekçesi / Analiz"]
            ws_analysis.append(headers)
            
            for cell in ws_analysis[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border

            for idx, log in enumerate(self.logs, 1):
                full_details = str(log["details"])
                selected_product = "-"
                reason = full_details
                
                if "Seçilen:" in full_details and "Karar:" in full_details:
                    parts = full_details.split("Karar:")
                    reason = parts[1].strip()
                    selected_product = parts[0].split("Seçilen:")[1].strip()
                elif "atandı" in full_details:
                    if "makinesine" in full_details:
                        parts = full_details.split("atandı.")
                        selected_product = parts[0].split("makinesine")[-1].strip()
                        reason = parts[1].strip() if len(parts) > 1 else full_details
                    else:
                        selected_product = full_details.split(" ")[0]
                        reason = full_details
                elif "planlandı" in full_details:
                    parts = full_details.split("planlandı.")
                    selected_product = parts[0].strip()
                    reason = parts[1].strip() if len(parts) > 1 else "Planlandı"
                elif "oluşturuldu" in full_details:
                    selected_product = "Toplu İş"
                    reason = full_details

                ws_analysis.append([idx, log["timestamp"].strftime('%d.%m.%Y %H:%M'), log["step"], log["action"], selected_product, reason])
                
                row_idx = idx + 1
                for col_idx in range(1, 7):
                    cell = ws_analysis.cell(row=row_idx, column=col_idx)
                    cell.alignment = left_align if col_idx >= 4 else center_align
                    cell.border = border
                    if log["step"] == "ASSEMBLY": cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            ws_analysis.column_dimensions['A'].width = 8
            ws_analysis.column_dimensions['B'].width = 18
            ws_analysis.column_dimensions['C'].width = 12
            ws_analysis.column_dimensions['D'].width = 25
            ws_analysis.column_dimensions['E'].width = 20
            ws_analysis.column_dimensions['F'].width = 100
            ws_analysis.auto_filter.ref = ws_analysis.dimensions
            
            wb.save(file_path)
        except Exception as e:
            print(f"Excel raporu oluşturulamadı: {e}")

    def generate_report_md(self, result_metadata: Dict[str, Any]) -> str:
        md = []
        md.append("# 📊 Algoritma Planlama ve Karar Analiz Raporu")
        md.append(f"**Oluşturulma Tarihi:** {datetime.now().strftime('%d.%m.%Y %H:%M')}")
        md.append(f"**Planlama Dönemi:** {result_metadata.get('period', 'N/A')}")
        md.append(f"**Toplam Üretilen Parça:** {result_metadata.get('total_parts', 0)}")
        md.append(f"**Toplam Setup Süresi:** {round(result_metadata.get('total_setup_time', 0), 2)} saat\n")
        
        md.append("## 🧠 Algoritma Yöntemi ve Karar Kriterleri")
        md.append("Bu planlama, **Öncelik Tabanlı Olay-Güdümlü (Priority-Based Event-Driven)** bir algoritma kullanır. Kararlar aşağıdaki kriterlere göre verilir:")
        md.append("1. **Excel Önceliği:** Excel'den gelen ürünler ve devam eden işler en yüksek öncelikle ele alınır.")
        md.append("2. **Teslimat Tarihi:** Vadesi yaklaşan işlerin (target_week) öncelik puanı otomatik yükseltilir.")
        md.append("3. **Üretim Akışı:** İstasyonlar arası parça transferi gerçek zamanlı olay kuyruğu ile yönetilir.")
        md.append("4. **Kapasite Verimliliği:** Paketleme (Assembly, B/N) aşamalarında makineler tam kapasite dolana kadar beklenir.\n")

        md.append("## ⚙️ Karar Mekanizması ve Detaylı Adımlar\n")
        
        # Olayları zaman sırasına göre grupla
        current_step = None
        for log in self.logs:
            if log["step"] != current_step:
                current_step = log["step"]
                md.append(f"### 📍 Adım: {current_step}")
            
            t_str = log["timestamp"].strftime('%Y-%m-%d %H:%M')
            md.append(f"- **[{t_str}]** {log['action']}: {log['details']}")
            
        md.append("\n--- \n*Bu rapor, planlama süreci boyunca algoritmanın verdiği kararları şeffaf bir şekilde izlemek için otomatik oluşturulmuştur.*")
        return "\n".join(md)
