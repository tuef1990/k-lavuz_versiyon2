from models.planning_result import PlanningResult
from services.metrics_service import MetricsService
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

class ExcelExportService:
    @staticmethod
    def export(result: PlanningResult, filepath: str):
        wb = openpyxl.Workbook()
        
        # Define common styles
        header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # ----------------------------------------------------
        # 1. Üretim Çizelgesi
        # ----------------------------------------------------
        ws1 = wb.active
        ws1.title = "Üretim Çizelgesi"
        
        headers = ["Ürün Tipi", "Ürün Adı", "İşlem Adı", "Makine", 
                   "Başlangıç Tarihi", "Başlangıç Saati", "Bitiş Tarihi", "Bitiş Saati", 
                   "Setup (saat)", "İşlem (saat)", "Kalan Süre (saat)", "Grup Adedi", "Vardiya", "Öncelik Puanı"]
        ws1.append(headers)
        
        # Style headers
        for col in range(1, len(headers) + 1):
            cell = ws1.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            
        # Add data
        row_colors = {
            "K11": "D6E4FF",
            "K12": "D1FAE5",
            "K13": "FEF3C7",
            "K14": "FEE2E2",
            "K15": "EDE9FE",
            "K16": "CFFAFE"
        }
        
        # Kalan süre: sadece üretim (process) süreleri toplanır, setup dahil edilmez.
        # Setup süresi parça tipine göre değişkenlik gösterdiğinden toplama katılırsa
        # farklı işler için farklı başlangıç değerleri ortaya çıkar.
        job_remaining_times = {}
        for e in result.schedule:
            if e.job_id not in job_remaining_times:
                job_remaining_times[e.job_id] = 0.0
            job_remaining_times[e.job_id] += e.process_time

        for e in sorted(result.schedule, key=lambda x: x.start_time):
            p_type = e.product_type
            p_name = e.product_name
            prio = e.priority_level

            dur = e.process_time

            # Bu adımın üretim süresi kadar düş
            job_remaining_times[e.job_id] -= e.process_time

            # Kayan nokta hatalarına karşı sıfırla
            if job_remaining_times[e.job_id] < 0.01:
                job_remaining_times[e.job_id] = 0.0

            # Bu adım bittikten sonra kalan süre
            current_remaining = job_remaining_times[e.job_id]
            
            row = [
                p_type, p_name, e.step_name, e.machine_name,
                e.start_time.strftime("%d.%m.%Y"),
                e.start_time.strftime("%H:%M"),
                e.end_time.strftime("%d.%m.%Y"),
                e.end_time.strftime("%H:%M"),
                e.setup_time, dur, round(current_remaining, 2),
                e.group_size,
                e.shift_number, round(prio, 3)
            ]
            ws1.append(row)
            
            # Color coding by product
            current_row = ws1.max_row
            hex_color = row_colors.get(p_type, "FFFFFF")
            fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            
            for col in range(1, len(headers) + 1):
                cell = ws1.cell(row=current_row, column=col)
                cell.fill = fill
                cell.border = border
                
        # Auto-fit columns
        for col in ws1.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws1.column_dimensions[column].width = adjusted_width
            
        ws1.auto_filter.ref = ws1.dimensions
        ws1.freeze_panes = "A2"

        # ----------------------------------------------------
        # 2. Performans Metrikleri
        # ----------------------------------------------------
        metrics = MetricsService.calculate(result)
        ws2 = wb.create_sheet(title="Performans Metrikleri")
        
        ws2.append(["Metrik", "Değer"])
        ws2.cell(row=1, column=1).font = header_font
        ws2.cell(row=1, column=1).fill = header_fill
        ws2.cell(row=1, column=2).font = header_font
        ws2.cell(row=1, column=2).fill = header_fill
        
        # Temel metrikler
        rows_basic = [
            ("Toplam Çizelge Süresi (Makespan)",       metrics["makespan_display"]),
            ("Son Parça Bitiş Zamanı",                  metrics["last_part_display"]),
            ("Toplam Üretilen Parça",                   metrics["total_parts"]),
            ("Günlük Verim (parça/gün)",                round(metrics["throughput_per_day"], 2)),
            ("Ort. Akış Süresi (saat/parça)",           metrics["avg_flow_time_hours"]),
            ("Hedef Karşılama Oranı (%)",               round(metrics["schedule_achievement_pct"], 1)),
        ]
        for label, value in rows_basic:
            ws2.append([label, value])

        ws2.append([])

        # Setup metrikleri
        rows_setup = [
            ("Toplam Setup Süresi (saat)",              round(metrics["total_setup_hours"], 2)),
            ("Toplam Setup İşlemi Sayısı",              metrics["num_setups"]),
            ("Setup Oranı (%)",                         round(metrics["setup_ratio_pct"], 2)),
        ]
        for label, value in rows_setup:
            ws2.append([label, value])

        ws2.append([])

        # Makine metrikleri
        rows_machine = [
            ("Ort. Makine Verimliliği (%)",             round(metrics["avg_utilization"], 2)),
            ("Darboğaz Makine",                         metrics["bottleneck_machine"]),
            ("Denge Endeksi (std sapma, düşük=iyi)",    metrics["balance_index"]),
        ]
        for label, value in rows_machine:
            ws2.append([label, value])

        ws2.append([])
        ws2.append(["Makine Verimliliği", "%"])
        r = ws2.max_row
        ws2.cell(row=r, column=1).font = header_font
        ws2.cell(row=r, column=1).fill = header_fill
        ws2.cell(row=r, column=2).font = header_font
        ws2.cell(row=r, column=2).fill = header_fill

        for m, util in metrics["machine_utilization"].items():
            idle = metrics["machine_idle_hours"].get(m, 0.0)
            ws2.append([m, util, idle])
            r2 = ws2.max_row
            cell = ws2.cell(row=r2, column=2)
            cell.number_format = '0.00"%"'
            if util > 80:   cell.fill = PatternFill(start_color="C6F6D5", fill_type="solid")
            elif util < 50: cell.fill = PatternFill(start_color="FED7D7", fill_type="solid")
            else:           cell.fill = PatternFill(start_color="FEEBC8", fill_type="solid")

        # Makine boşta süresi başlığını geriye dönük ekle
        util_header_row = r
        ws2.cell(row=util_header_row, column=3).value = "Boşta Süre (saat)"
        ws2.cell(row=util_header_row, column=3).font = header_font
        ws2.cell(row=util_header_row, column=3).fill = header_fill

        ws2.column_dimensions['A'].width = 40
        ws2.column_dimensions['B'].width = 20
        ws2.column_dimensions['C'].width = 20
        
        # ----------------------------------------------------
        # Makine Detayı
        # ----------------------------------------------------
        ws4 = wb.create_sheet(title="Makine Detayı")
        
        # Group entries by machine
        by_machine = {}
        for e in result.schedule:
            if e.machine_name not in by_machine: by_machine[e.machine_name] = []
            by_machine[e.machine_name].append(e)
            
        current_row = 1
        for m in sorted(by_machine.keys()):
            ws4.cell(row=current_row, column=1).value = f"Makine/İstasyon: {m}"
            ws4.cell(row=current_row, column=1).font = Font(size=14, bold=True)
            current_row += 1
            
            ws4.append(["Ürün Adı", "İşlem", "Baş. Tarih", "Baş. Saat", "Bitiş Tarih", "Bitiş Saat", "Süre (saat)", "Adet"])
            for col in range(1, 9):
                c = ws4.cell(row=current_row, column=col)
                c.fill = PatternFill(start_color="E2E8F0", fill_type="solid")
                c.font = Font(bold=True)
            current_row += 1
            
            entries = sorted(by_machine[m], key=lambda x: x.start_time)
            from datetime import timedelta
            for e in entries:
                if e.setup_time > 0:
                    setup_end = e.start_time + timedelta(hours=e.setup_time)
                    ws4.append(["SETUP", e.step_name, e.start_time.strftime("%d.%m.%Y"), e.start_time.strftime("%H:%M"), 
                                setup_end.strftime("%d.%m.%Y"), setup_end.strftime("%H:%M"), e.setup_time, "-"])
                    for col in range(1, 9):
                        ws4.cell(row=current_row, column=col).font = Font(italic=True, color="718096")
                    current_row += 1
                    process_start = setup_end
                else:
                    process_start = e.start_time
                    
                if e.process_time > 0:
                    p_name = e.product_name
                    ws4.append([p_name, e.step_name, process_start.strftime("%d.%m.%Y"), process_start.strftime("%H:%M"),
                                e.end_time.strftime("%d.%m.%Y"), e.end_time.strftime("%H:%M"), e.process_time, e.group_size])
                    current_row += 1
                
            current_row += 2 # Spacer

        ws4.column_dimensions['A'].width = 15
        ws4.column_dimensions['B'].width = 15
        ws4.column_dimensions['C'].width = 18
        ws4.column_dimensions['D'].width = 18
        ws4.column_dimensions['E'].width = 10
        ws4.column_dimensions['F'].width = 10

        # ----------------------------------------------------
        # 5. Grup Üretim Akışı
        #    Assembly batch = grup. Her grup için tüm adımlar (Assembly,
        #    FTP, B/N, DKK, RVB, ATP+STP) sırayla listelenir, sonra
        #    bir sonraki gruba geçilir.
        # ----------------------------------------------------
        ws5 = wb.create_sheet(title="Grup Üretim Akışı")
        h5 = [
            "Grup No", "Ürün Tipi", "Ürün Adı", "İşlem Türü",
            "Başlangıç", "Bitiş", "Süre (saat)", "Makine",
        ]
        ws5.append(h5)
        for col in range(1, len(h5) + 1):
            cell = ws5.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        # Step-batch bazlı gruplama: aynı (start_time, makine, adım) anahtarına
        # düşen tüm parçalar tek bir step-batch'tir. Her step-batch zamana göre
        # sıralanır ve 1'den başlayarak numaralandırılır — böylece tablodaki
        # satırlar zaman akışına göre sıralı 1, 2, 3, 4 ... şeklinde gözükür.
        from collections import defaultdict
        step_batches: dict = defaultdict(list)
        for e in result.schedule:
            key = (e.start_time, e.machine_name, e.step_name.lower())
            step_batches[key].append(e)

        # Adım label eşleştirmesi (UI için okunaklı)
        STEP_LABELS = {
            "assembly": "Assembly",
            "ftp": "FTP",
            "bn": "B/N",
            "dkk": "DKK",
            "rvb": "RVB",
            "atp_stp": "ATP+STP",
        }

        def _fmt_dt(dt):
            return dt.strftime("%d.%m.%Y %H:%M") if dt else "-"

        # Zaman sırasına göre dizilim (eşit start_time'da makine adıyla tie-break)
        sorted_batches = sorted(
            step_batches.items(),
            key=lambda kv: (kv[0][0], kv[0][1])
        )

        # Renkler:
        #   - FIRST_ROW_COLOR: her grubun ilk ürünü için belirgin vurgu (açık mavi)
        #   - REST_ROW_COLOR: aynı grubun kalan satırları (beyaz)
        FIRST_ROW_COLOR = "DBEAFE"   # açık mavi — grubun başlangıcını işaretler
        REST_ROW_COLOR  = "FFFFFF"   # beyaz
        first_fill = PatternFill(start_color=FIRST_ROW_COLOR, end_color=FIRST_ROW_COLOR, fill_type="solid")
        rest_fill  = PatternFill(start_color=REST_ROW_COLOR, end_color=REST_ROW_COLOR, fill_type="solid")
        first_font = Font(bold=True)

        for g_no, (key, entries) in enumerate(sorted_batches, start=1):
            _start, machine, step_key = key
            # Aynı batch içindeki parçaları ürün-adı ile sırala (deterministik)
            entries_sorted = sorted(entries, key=lambda e: (e.product_type, e.product_name, e.job_id))
            step_label = STEP_LABELS.get(step_key, step_key.upper())
            for row_idx, ent in enumerate(entries_sorted):
                duration = round((ent.end_time - ent.start_time).total_seconds() / 3600.0, 2)
                row = [
                    f"Grup {g_no}",
                    ent.product_type,
                    ent.product_name,
                    step_label,
                    _fmt_dt(ent.start_time),
                    _fmt_dt(ent.end_time),
                    duration,
                    ent.machine_name,
                ]
                ws5.append(row)
                cur = ws5.max_row
                # Grubun ilk satırı (row_idx == 0) vurgulanır; kalanlar düz beyaz
                is_first = row_idx == 0
                fill = first_fill if is_first else rest_fill
                for col in range(1, len(h5) + 1):
                    cell = ws5.cell(row=cur, column=col)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = Alignment(vertical="center", horizontal="center")
                    if is_first:
                        cell.font = first_font

        # Sütun genişlikleri
        widths = [10, 12, 14, 14, 18, 18, 12, 14]
        for i, w in enumerate(widths, start=1):
            ws5.column_dimensions[ws5.cell(row=1, column=i).column_letter].width = w
        ws5.auto_filter.ref = ws5.dimensions
        ws5.freeze_panes = "A2"

        wb.save(filepath)
        return True
