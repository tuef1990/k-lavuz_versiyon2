"""
Excel dosyasından ürün bilgilerini okuyarak planlama algoritmasına uygun formata çeviren servis.

Excel dosyasındaki sütunlar:
- Ürün Tipi, Ürün Adı, Adet
- İstasyon sütunları (Montaj, FTP, B/N, DKK, RVB, ATP+STB) → Hangi adımda olduğunu belirler
- Muhtemel başlatılabilme tarihi → earliest_start_date
- Kalan İşçilik → Mevcut adımdaki kalan süre (saat)
"""

from dataclasses import dataclass, field
from datetime import datetime
from typing import List, Optional, Dict, Tuple
import openpyxl
import os
import re


@dataclass
class ExcelProduct:
    """Excel'den okunan tek bir ürün satırını temsil eder."""
    product_type: str           # Ürün Tipi (K11, K12 vb.)
    product_name: str           # Ürün Adı
    quantity: int               # Adet / Miktar
    current_step: str           # Mevcut adım (assembly, ftp, bn, dkk, rvb, atp_stp)
    remaining_work_hours: float # Kalan İşçilik Süresi (saat)
    earliest_start_date: Optional[datetime]  # Muhtemel Başlatılabilme Tarihi
    completed_steps: List[str] = field(default_factory=list)  # Tamamlanmış adımlar
    preferred_machine: Optional[str] = None  # Atanmış makine (M1/M2/M3/M4), Excel'den gelir
    period_target: Optional[int] = None  # Kullanıcının elle girdiği periyot hedefi (None → quantity kullanılır)
    # Aynı satırdan virgülle ayrılarak gelen ürünler aynı group_id'yi paylaşır.
    # Planner'lar batch oluştururken bu işleri tek grup olarak ele alır.
    group_id: Optional[str] = None
    # Tablolar (üretim süresi, kapasite, setup) tip başına TEK satır tutar.
    # Aynı tipten birden fazla isim Excel'de varsa (örn K12/A, K12/B), hepsi
    # canonical_pid (=tablodaki tek girişin display_name'i) üzerinden okunur.
    # Default: ürünün kendi display_name'i (kendi tipinin canonical'ı).
    canonical_pid: Optional[str] = None

    @property
    def display_name(self) -> str:
        return f"{self.product_type} / {self.product_name}"


# İstasyon sütun isimlerini algoritmanın kullandığı step isimlerine eşleyen harita
# Excel'deki sütun isimleri çeşitli formatlarda gelebilir, bu yüzden esnek matching yapıyoruz
# Makine adı tanıma pattern'i (M1, M2, M3, M4)
_MACHINE_PATTERN = re.compile(r'\b(M[1-4])\b', re.IGNORECASE)

STATION_COLUMN_MAPPING = {
    # Excel sütun adı pattern'leri -> algoritma step adı
    'montaj': 'assembly',
    'assembly': 'assembly',
    'ftp': 'ftp',
    'b/n': 'bn',
    'bn': 'bn',
    'b\\n': 'bn',
    'dkk': 'dkk',
    'ddk': 'dkk',      # Olası yazım hatası
    'rvb': 'rvb',
    'atp+stb': 'atp_stp',
    'atp+stp': 'atp_stp',
    'atp': 'atp_stp',
    'stp': 'atp_stp',
}

# Adım sıralaması (tamamlanmış adımları belirlemek için)
STEP_ORDER = ['assembly', 'ftp', 'bn', 'dkk', 'rvb', 'atp_stp']


def _normalize_column_name(name: str) -> str:
    """Sütun adını normalize eder (küçük harf, boşluk/parantez temizle)."""
    if not name:
        return ""
    # Python'daki Türkçe I/İ lower() hatasını önlemek için:
    name = name.replace('İ', 'i').replace('I', 'ı')
    return name.strip().lower()


def _match_station_column(col_name: str) -> Optional[str]:
    """
    Bir sütun adının istasyon sütunu olup olmadığını kontrol eder.
    Eğer eşleşme varsa algoritma step adını döner ('assembly', 'ftp' vb.)
    """
    normalized = _normalize_column_name(col_name)
    
    # Doğrudan eşleşme
    if normalized in STATION_COLUMN_MAPPING:
        return STATION_COLUMN_MAPPING[normalized]
    
    # Parantez ve ekstra bilgileri kaldırarak eşleştir
    # Örn: "RVB (X-Y/Z)" -> "rvb"
    base_name = re.split(r'[\s\(\[]', normalized)[0].strip()
    if base_name in STATION_COLUMN_MAPPING:
        return STATION_COLUMN_MAPPING[base_name]
    
    return None


def _get_completed_steps(current_step: str) -> List[str]:
    """Mevcut adımdan önceki tüm adımları tamamlanmış olarak döner."""
    if current_step not in STEP_ORDER:
        return []
    
    current_idx = STEP_ORDER.index(current_step)
    return STEP_ORDER[:current_idx]


def _parse_date(value) -> Optional[datetime]:
    """Tarih değerini parse eder. Çeşitli formatları destekler."""
    if value is None or str(value).strip() == "":
        return None
    
    if isinstance(value, datetime):
        return value
    
    # String tarih formatlarını dene
    date_str = str(value).strip()
    formats = [
        "%d.%m.%Y", "%d.%m.%Y %H:%M",
        "%Y-%m-%d", "%Y-%m-%d %H:%M",
        "%d/%m/%Y", "%d/%m/%Y %H:%M",
        "%m/%d/%Y", "%m/%d/%Y %H:%M",
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    
    return None


def _parse_float(value) -> float:
    """Sayısal değeri float olarak parse eder."""
    if value is None or str(value).strip() == "":
        return 0.0
    
    try:
        return float(str(value).replace(",", "."))
    except (ValueError, TypeError):
        return 0.0


def _parse_int(value) -> int:
    """Sayısal değeri int olarak parse eder."""
    if value is None or str(value).strip() == "":
        return 0
    
    try:
        return int(float(str(value).replace(",", ".")))
    except (ValueError, TypeError):
        return 0


class ExcelImportService:
    """Excel dosyasını okuyup ExcelProduct listesi oluşturan servis."""
    
    @staticmethod
    def read_excel(file_path: str) -> Tuple[List[ExcelProduct], List[str]]:
        """
        Excel dosyasını okur ve ExcelProduct listesi döner.
        
        Args:
            file_path: Excel dosyasının yolu
            
        Returns:
            Tuple of (ürün listesi, hata/uyarı mesajları listesi)
        """
        products: List[ExcelProduct] = []
        warnings: List[str] = []
        
        if not os.path.exists(file_path):
            return [], [f"Dosya bulunamadı: {file_path}"]
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
        except Exception as e:
            return [], [f"Excel dosyası açılamadı: {str(e)}"]
        
        # Başlık satırını bul: önce 1. satıra bak; sütun isimleri bulunamazsa 2. satıra düş.
        # Hangi satır seçilirse veri o satırın bir sonrasından okunmaya başlar.
        def _read_row(row_idx: int) -> List[str]:
            return [str(cell.value) if cell.value else "" for cell in ws[row_idx]]

        header_row = 1
        headers = _read_row(1)
        col_map = ExcelImportService._identify_columns(headers)

        # 1. satırda zorunlu sütunlar bulunamadıysa 2. satıra bak
        if col_map['type_col'] is None or col_map['name_col'] is None:
            try:
                row2 = _read_row(2)
                col_map_row2 = ExcelImportService._identify_columns(row2)
                if (col_map_row2['type_col'] is not None
                        and col_map_row2['name_col'] is not None):
                    header_row = 2
                    headers = row2
                    col_map = col_map_row2
                    warnings.append("ℹ️ İlk satırda sütun isimleri bulunamadı, 2. satır başlık olarak kullanıldı.")
            except Exception:
                pass

        if len(headers) < 3:
            return [], ["Excel dosyasında yeterli sütun bulunamadı. En az Ürün Tipi, Ürün Adı ve Adet sütunları gereklidir."]

        if col_map['type_col'] is None:
            warnings.append("⚠️ 'Ürün Tipi' sütunu bulunamadı.")
        if col_map['name_col'] is None:
            warnings.append("⚠️ 'Ürün Adı' sütunu bulunamadı.")
        if col_map['quantity_col'] is None:
            warnings.append("⚠️ 'Adet' sütunu bulunamadı.")

        if col_map['type_col'] is None or col_map['name_col'] is None:
            return [], warnings + ["Ürün Tipi ve Ürün Adı sütunları zorunludur."]

        # Veri satırlarını oku (header_row+1'den itibaren)
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue  # Boş satırı atla
            
            # Sütun sayısı header'dan az ise row'u pad'le
            row = list(row) + [None] * max(0, len(headers) - len(row))
            
            try:
                product_type = str(row[col_map['type_col']] or "").strip()
                raw_name = str(row[col_map['name_col']] or "").strip()

                if not product_type and not raw_name:
                    continue  # İkisi de boşsa satırı atla

                # Ürün Adı alanı virgülle ayrılmış birden fazla isim içerebilir.
                # Örn: "A, B, C" → 3 ayrı ExcelProduct olarak işle (her biri aynı tip + adet ile).
                name_list = [n.strip() for n in raw_name.split(",") if n.strip()]
                if not name_list:
                    name_list = [""]  # adı boş olsa da ürün tipi varsa tek satır olarak ele al

                quantity = _parse_int(row[col_map['quantity_col']]) if col_map['quantity_col'] is not None else 1

                # İstasyon sütunlarından mevcut adımı ve makine atamasını belirle
                current_step = 'assembly'  # Varsayılan: baştan başla
                preferred_machine = None
                found_step = False

                for station_col_idx, step_name in col_map['station_cols']:
                    cell_value = row[station_col_idx] if station_col_idx < len(row) else None
                    if cell_value is not None and str(cell_value).strip() != "":
                        current_step = step_name
                        found_step = True
                        # Hücre değeri makine adı içeriyorsa (M1/M2/M3/M4) kaydet
                        m = _MACHINE_PATTERN.search(str(cell_value))
                        if m:
                            preferred_machine = m.group(1).upper()
                        break

                if not found_step:
                    warnings.append(f"Satır {row_idx}: '{product_type} / {raw_name}' için istasyon bilgisi bulunamadı, Assembly olarak varsayıldı.")

                # Tamamlanmış adımları belirle
                completed_steps = _get_completed_steps(current_step)

                # Kalan İşçilik Süresi
                remaining_hours = _parse_float(row[col_map['remaining_col']]) if col_map['remaining_col'] is not None else 0.0

                # Muhtemel Başlatılabilme Tarihi
                earliest_date = _parse_date(row[col_map['date_col']]) if col_map['date_col'] is not None else None

                # Aynı satırdan virgülle ayrılan ürünler ortak group_id paylaşır
                # → planner'lar bunları batch olarak işler. Tek isimliyse group_id None.
                row_group_id = (
                    f"row{row_idx}_{product_type}" if len(name_list) > 1 else None
                )

                # Adet dağıtımı: Adet sütunu TOPLAM olarak yorumlanır.
                # Birden fazla isim varsa eşit dağıtılır; bölünemiyorsa kalan ilk isimlere +1.
                # Örn: "A,B,C,D,E, 25" → her birine 5
                #      "A,B,C, 10"     → 4, 3, 3
                if len(name_list) > 1:
                    base = quantity // len(name_list)
                    remainder = quantity - base * len(name_list)
                    qty_per_name = [base + (1 if i < remainder else 0) for i in range(len(name_list))]
                else:
                    qty_per_name = [quantity]

                # Her isim için ayrı ExcelProduct oluştur (aynı tip, adım, vs.) — adet dağıtılır
                for idx_n, product_name in enumerate(name_list):
                    product = ExcelProduct(
                        product_type=product_type,
                        product_name=product_name,
                        quantity=qty_per_name[idx_n],
                        current_step=current_step,
                        remaining_work_hours=remaining_hours,
                        earliest_start_date=earliest_date,
                        completed_steps=list(completed_steps),
                        preferred_machine=preferred_machine,
                        group_id=row_group_id,
                    )
                    products.append(product)

                if len(name_list) > 1:
                    distribution = ", ".join(f"{n}={q}" for n, q in zip(name_list, qty_per_name))
                    warnings.append(
                        f"Satır {row_idx}: '{product_type}' tipi için {len(name_list)} farklı isim "
                        f"({', '.join(name_list)}) AYNI grup ({row_group_id}) olarak okundu. "
                        f"Toplam {quantity} adet eşit dağıtıldı: {distribution}."
                    )

            except Exception as e:
                warnings.append(f"Satır {row_idx}: Okuma hatası - {str(e)}")
        
        wb.close()
        
        if not products:
            warnings.append("Excel dosyasından hiç ürün okunamadı.")
        else:
            warnings.insert(0, f"✅ {len(products)} ürün başarıyla okundu.")
        
        return products, warnings
    
    @staticmethod
    def _identify_columns(headers: List[str]) -> Dict:
        """
        Header satırından sütun indekslerini belirler.
        Esnek matching kullanır: normalize edilmiş isimleri karşılaştırır.
        """
        col_map = {
            'type_col': None,      # Ürün Tipi
            'name_col': None,      # Ürün Adı
            'quantity_col': None,   # Adet / Miktar
            'remaining_col': None,  # Kalan İşçilik Süresi
            'date_col': None,       # Muhtemel Başlatılabilme Tarihi
            'station_cols': [],     # [(col_idx, step_name), ...]
        }
        
        for idx, header in enumerate(headers):
            normalized = _normalize_column_name(header)
            
            if not normalized:
                continue
            
            # Ürün Tipi
            if any(k in normalized for k in ['ürün tipi', 'urun tipi', 'tip', 'type']):
                col_map['type_col'] = idx
                continue
            
            # Ürün Adı
            if any(k in normalized for k in ['ürün adı', 'urun adi', 'ürün ad', 'product name']):
                col_map['name_col'] = idx
                continue
            
            # Adet / Miktar
            if any(k in normalized for k in ['adet', 'miktar', 'quantity', 'qty']):
                col_map['quantity_col'] = idx
                continue
            
            # Kalan İşçilik
            if any(k in normalized for k in ['kalan işçilik', 'kalan iscilik', 'kalan süre', 'remaining']):
                col_map['remaining_col'] = idx
                continue
            
            # Muhtemel Başlatılabilme Tarihi
            if any(k in normalized for k in ['muhtemel', 'başlatılabilme', 'baslatilabilme', 'tarih', 'start date']):
                col_map['date_col'] = idx
                continue
            
            # İstasyon sütunlarını kontrol et
            step_name = _match_station_column(header)
            if step_name:
                col_map['station_cols'].append((idx, step_name))
        
        # İstasyon sütunlarını STEP_ORDER sırasına göre sırala
        col_map['station_cols'].sort(key=lambda x: STEP_ORDER.index(x[1]) if x[1] in STEP_ORDER else 999)
        
        return col_map
    
    @staticmethod
    def validate_against_app_data(products: List[ExcelProduct], app_product_names: List[str]) -> List[str]:
        """
        Excel'den okunan ürünlerin app'teki mevcut ürünlerle eşleşip eşleşmediğini kontrol eder.
        Eşleşmeyen ürünler için uyarı verir (kapasite, süre vb. bilgisi olmayabilir).
        
        Args:
            products: Excel'den okunan ürünler
            app_product_names: App'teki mevcut ürün display_name listesi
            
        Returns:
            Uyarı mesajları listesi
        """
        warnings = []
        
        for ep in products:
            if ep.display_name not in app_product_names:
                warnings.append(
                    f"⚠️ '{ep.display_name}' ürünü uygulamadaki tablolarda tanımlı değil. "
                    f"Kapasite ve üretim süresi bilgileri bulunamayabilir."
                )
        
        return warnings

    @staticmethod
    def read_raw_rows(file_path: str):
        """
        Excel dosyasını açar ve tüm satırları (header dahil) ham liste olarak döndürür.
        Her satır bir liste; her eleman o hücrenin değeri (None olabilir).
        İlk liste (index 0) başlık (header) satırıdır.
        
        Returns:
            List[List]: [[header1, header2, ...], [val1, val2, ...], ...]
        """
        if not os.path.exists(file_path):
            return []
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            wb.close()
            return rows
        except Exception:
            return []

